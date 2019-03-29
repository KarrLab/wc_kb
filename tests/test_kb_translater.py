from wc_kb import kb_translater
import os
import shutil
import tempfile
import unittest
import openpyxl

class TestKbTranslater(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.translator = kb_translater.KbTranslater(
            kb_path = 'kbs/kb_core_empty.xlsx',
            core_path = 'kbs/original_core.xlsx')

    def test_delinateJSONs(self):

        assert kb_translater.KbTranslater.delinateJSONs('') == None
        assert kb_translater.KbTranslater.delinateJSONs(None) == None

        assert kb_translater.KbTranslater.delinateJSONs('{"a":"b"}') == [{'a':'b'}]
        assert kb_translater.KbTranslater.delinateJSONs('{"a":"b"}, {"A":1,"B":2,"C":3}') == [{"a":"b"}, {"A":1,"B":2,"C":3}]

        with self.assertWarns(Warning):
            assert kb_translater.KbTranslater.delinateJSONs('[4,5,6]') is None
            assert kb_translater.KbTranslater.delinateJSONs('{"a":"b"}, {"A":1,"B":2,"C":3}, [4,5,6]') == [{"a":"b"}, {"A":1,"B":2,"C":3}]
            assert kb_translater.KbTranslater.delinateJSONs('[{"a":"b"}, {"A":1,"B":2,"C":3}]') == [{"a":"b"}, {"A":1,"B":2,"C":3}]
            assert kb_translater.KbTranslater.delinateJSONs('[{"a":"b"}, {"A":1,"B":2,"C":3}, [4,5,6]]') == [{"a":"b"}, {"A":1,"B":2,"C":3}]

        with self.assertRaises(Exception):
            kb_translater.KbTranslater.delinateJSONs('{testing}')

        # Can not parse JSONs from line: {"test":{"a":1, "b":2}} #Nested case

    def test_concatenateValues(self):

        assert kb_translater.KbTranslater.concatenateValues(None, 'a') is None
        assert kb_translater.KbTranslater.concatenateValues([{"a":1}], 'a') == '1'
        assert kb_translater.KbTranslater.concatenateValues([{"a":1}, {"a":11,"b":2,"c":3}], 'a') == '1, 11'
        assert kb_translater.KbTranslater.concatenateValues([{"a":'1'}, {"a":'11',"b":2,"c":3}], 'a') == '1, 11'

        assert kb_translater.KbTranslater.concatenateValues([{"a":1, "A":2}, {"a":11,"b":2,"c":3,"A":3}], ('a', 'A')) == '1:2, 11:3'
        assert kb_translater.KbTranslater.concatenateValues([{"a":1, "A":"A"}, {"a":11,"b":2,"c":3,"A":"B"}], ('a', 'A')) == '1:A, 11:B'

        with self.assertRaises(AssertionError):
            kb_translater.KbTranslater.concatenateValues({"a":1}, 'a') == '1'

        with self.assertRaises(KeyError):
            kb_translater.KbTranslater.concatenateValues([{"a":1, "A":2}, {"a":11,"b":2,"c":3}], ('a', 'A')) == '1:2, 11:3'

    def test_concatenateEviField(self):

        wb = openpyxl.Workbook()
        wb['Sheet'].cell(column=1, row=1).value = 'test1'
        wb['Sheet'].cell(column=2, row=1).value = None

        # Testing None case
        kb_translater.KbTranslater.concatenateEviField(wb['Sheet'].cell(column=2, row=1), 'empty line test')
        assert wb['Sheet'].cell(column=2, row=1).value == 'empty line test'

        # Testing case with pre-existing entry
        kb_translater.KbTranslater.concatenateEviField(wb['Sheet'].cell(column=1, row=1), 'test2')
        assert wb['Sheet'].cell(column=1, row=1).value == 'test1, test2'

        kb_translater.KbTranslater.concatenateEviField(wb['Sheet'].cell(column=2, row=1), 'test3')
        assert wb['Sheet'].cell(column=2, row=1).value == 'empty line test, test3'

    def test_getColumnId(self):

        with self.assertRaises(Exception):
            self.translator.getColumnId(wbName = 'test', sheetName = 'Genes', header = 'Id')

        self.assertEqual(self.translator.getColumnId(wbName='kb', sheetName='Genes', header = 'ID'), 1)
        self.assertEqual(self.translator.getColumnId(wbName='kb', sheetName='Genes', header = 'id'), 1)
        self.assertEqual(self.translator.getColumnId(wbName='kb', sheetName='Genes', header = 'Comments'), 16)
        self.assertEqual(self.translator.getColumnId(wbName='kb', sheetName='Species properties', header = 'Name'), 2)
        self.assertEqual(self.translator.getColumnId(wbName='kb', sheetName='Species properties', header = 'Half life'), 5)

        self.assertEqual(self.translator.getColumnId(wbName='core', sheetName='Genes', header = 'cross references'), 5)
        self.assertEqual(self.translator.getColumnId(wbName='core', sheetName='Genes', header = 'Cross references'), 5)
        self.assertEqual(self.translator.getColumnId(wbName='core', sheetName='Compartments', header = 'NAME'), 2)
        self.assertEqual(self.translator.getColumnId(wbName='core', sheetName='Protein monomers', header = 'Evidence'), [8, 13, 19, 23, 26, 29, 32])
