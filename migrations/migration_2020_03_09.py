""" Migration to ObjTables format as of 2020-03-09

:Author: Jonathan Karr <karr@mssm.edu>
:Date: 2020-03-09
:Copyright: 2020, Karr Lab
:License: MIT
"""

import wc_kb
import wc_kb.eukaryote
import wc_kb.prokaryote
import openpyxl
import re
import stringcase


def transform(filename, taxon):
    # read
    wb = openpyxl.load_workbook(filename=filename)

    for ws in wb:
        if not ws.title.startswith('!'):
            continue

        # lower camel case document and class attributes
        table_head_cell = None
        if isinstance(ws.cell(1, 1).value, str) and ws.cell(1, 1).value.startswith('!!'):
            table_head_cell = ws.cell(1, 1)
            matches = re.findall(r" +(.*?)=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")",
                ws.cell(1, 1).value)
            heading, _, _ = ws.cell(1, 1).value.partition(' ')
            for key, val, _, _ in matches:
                heading += ' {}={}'.format(stringcase.camelcase(key), val)
            ws.cell(1, 1).value = heading

        if isinstance(ws.cell(2, 1).value, str) and ws.cell(2, 1).value.startswith('!!'):
            table_head_cell = ws.cell(2, 1)
            matches = re.findall(r" +(.*?)=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")",
                ws.cell(2, 1).value)
            heading, _, _ = ws.cell(2, 1).value.partition(' ')
            for key, val, _, _ in matches:
                heading += ' {}={}'.format(stringcase.camelcase(key), val)
            ws.cell(2, 1).value = heading

        # set schema
        if taxon == 'eu':
            schema = 'wc_kb.eukaryote'
        else:
            schema = 'wc_kb.prokaryote'

        if ws.title == '!!_Schema':
            raise NotImplementedError('setting schema name not supported')
        elif ws.title != '!!_Table of contents':
            table_head_cell.value += " schema='{}'".format(schema)

        # set table format
        if ws.title in ['!!_Schema', '!!_Table of contents']:
            table_head_cell.value += ' tableFormat="row"'
        else:
            match = re.search(r" +id=('((?:[^'\\]|\\.)*)'|\"((?:[^\"\\]|\\.)*)\")",
                table_head_cell.value)
            table_id = match.group(1)[1:-1]
            if hasattr(wc_kb, table_id):
                table = getattr(wc_kb, table_id)
            elif taxon == 'eu':
                table = getattr(wc_kb.eukaryote, table_id)
            else:
                table = getattr(wc_kb.prokaryote, table_id)

            table_format = table.Meta.table_format.name
            table_head_cell.value += ' tableFormat="{}"'.format(table_format)

    # save
    wb.save(filename)
