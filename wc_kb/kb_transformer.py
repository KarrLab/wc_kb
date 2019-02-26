""" Parsing of the core.xlsx unstructured data file and insering fields to kb_core.xlsx

:Author: Balazs Szigeti <balazs.szigeti@mssm.edu>
:Date: 2019-02-23
:Copyright: 2019, Karr Lab
:License: MIT
"""

import wc_kb
import json
import openpyxl
import pdb
import re

class KbTransformer:

    def __init__(self, core=None, kb=None):
        if core is None:
            self.core = openpyxl.load_workbook('kbs/core.original.xlsx')
        if kb is None:
            self.kb = openpyxl.load_workbook('kbs/kb_core.xlsx')

    def transcodeReferences(self):
        """ Parses information from the References's 'Cross references' column and inserts it to the appropiate field(s) in the new kb structure. """

        coreRefs = self.core['References']
        kbRefs   = self.kb['References']
        cell2str = lambda myCell: myCell or "" # Convert None to string or return original string

        for row_idx in range(3, coreRefs.max_row+1):

            # Make sure IDs are matching; skip if there is no database ref
            assert(kbRefs.cell(column=1, row=row_idx-1).value == coreRefs.cell(column=1, row=row_idx).value)
            coreDbref = coreRefs.cell(column=4, row=row_idx).value
            if coreDbref is None:
                continue

            # Get list of JSONs in string add them to DBrefs / comments
            jsons = self.parseJsonStrs(coreDbref)
            for json in jsons: # Each JSON is a dict with a single key (DB name) - value pair (xid)
                if json['source'] == 'URL':
                    myCell = kbRefs.cell(column=12, row=row_idx-1).value
                    kbRefs.cell(column=12, row=row_idx-1).value = cell2str(myCell) + str(json['xid']) + ', '
                else:
                    myCell = kbRefs.cell(column=11, row=row_idx-1).value
                    kbRefs.cell(column=11, row=row_idx-1).value = cell2str(myCell) + "{}:{}".format(json['source'], json['xid']) + ', '
                    self.addDatabaseReference(json)

            # Remove " ," fromt end of DBrefs / comments
            if kbRefs.cell(column=11, row=row_idx-1).value is not None:
                kbRefs.cell(column=11, row=row_idx-1).value  = kbRefs.cell(column=11, row=row_idx-1).value[:-2]

            if kbRefs.cell(column=12, row=row_idx-1).value is not None:
                kbRefs.cell(column=12, row=row_idx-1).value  = kbRefs.cell(column=12, row=row_idx-1).value[:-2]

        self.kb.save('kbs/kb_core2.xlsx')

    def addDatabaseReference(self, json):
        wsName = 'Database references'
        kbDbRefs = self.kb[wsName]
        nextRow  = self.kb[wsName].max_row+1 #self.nextRow(wsName)
        dbRefId = "{}:{}".format(json['source'], json['xid'])

        if not self.idExists(wsName, dbRefId):
            kbDbRefs.cell(column=1, row=nextRow).value = "{}:{}".format(json['source'], json['xid'])
            kbDbRefs.cell(column=2, row=nextRow).value = json['source']
            kbDbRefs.cell(column=3, row=nextRow).value = str(json['xid'])

    def idExists(self, wsName, id):
        sheet = self.kb[wsName]
        nextRow = self.kb[wsName].max_row+1 #self.nextRow(wsName)

        for row in range(1, nextRow):
            if sheet.cell(column=1, row=row).value == id:
                return True

        return False

    @staticmethod
    def parseJsonStrs(fieldStr):
        """ Detects and return individual JSON strings within fieldStr """

        jsons = []
        for match in re.findall(r'{.*?}', fieldStr):
            try:
                jsons.append(json.loads(match))
            except:
                pdb.set_trace()
                raise Exception('Can not parse JSONs from line: {}'.format(fieldStr))

        return jsons


    """
    def nextRow(self, wsName):
        " Returns the index of the next empty row in worksheet with name wsName. ""

        for nextRowIdx, row in enumerate(self.kb[wsName], 1):
            print(nextRowIdx, row)
            if self.kb[wsName].cell(column=1, row=row).value is None:
                return nextRowIdx

            return (nextRowIdx+1)

        for nextRowIdx, row in enumerate(self.kb[wsName], 1):
            if all(c.value is None for c in row):
                return nextRowIdx

            nextRowIdx += 1
            return nextRowIdx
    """
