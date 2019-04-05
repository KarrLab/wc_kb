""" Parsing of the core.xlsx unstructured data file and insering fields to kb_core.xlsx

:Author: Balazs Szigeti <balazs.szigeti@mssm.edu>
:Date: 2019-02-23
:Copyright: 2019, Karr Lab
:License: MIT
"""

import wc_kb
import json
import openpyxl
import pdb

# Load workbooks and get sheets
core = openpyxl.load_workbook('kbs/core.original.xlsx')
core_refs = core['References']
kb_core = openpyxl.load_workbook('kbs/kb_core.xlsx')
kb_refs = kb_core['References']


for row_idx in range(3,core_refs.max_row+1):
    # Make sure IDs are matching;
    assert(kb_refs.cell(column=1, row=row_idx-1).value == core_refs.cell(column=1, row=row_idx).value)

    core_dbref = core_refs.cell(column=4, row=row_idx).value
    if core_dbref is None:     # Skip if there is no database ref
        continue

    # Check the number of JSON entries
    if '}, {' in core_dbref or '},{' in core_dbref:
        pdb.set_trace()
        continue

    try:
        temp = json.loads(core_refs.cell(column=4, row=row_idx).value)
    except:
        print('Row {}: {}'.format(row_idx, core_refs.cell(column=4, row=row_idx).value))
        continue

    database_ref_id = "{}:{}".format(temp['source'], temp['xid'])
    kb_refs.cell(column=11, row=row_idx-1).value = database_ref_id

kb_core.save('kb_core_ParseAdded.xlsx')
