""" Parsing of the core.xlsx unstructured data file and insering fields to kb_core.xlsx

:Author: Balazs Szigeti <balazs.szigeti@mssm.edu>
:Date: 2019-02-23
:Copyright: 2019, Karr Lab
:License: MIT
"""

import wc_kb
import json
import openpyxl
import warnings
import pdb
import re

class KbTranslater:

    def __init__(self, core_path=None, kb_path=None, column=None):

        if core_path is None:
            core_path = 'wc_kb/kbs/original_core.xlsx'
        if kb_path is None:
            kb_path = 'wc_kb/kbs/kb_core_empty.xlsx'

        self.core = openpyxl.load_workbook(core_path, read_only=True)
        self.kb = openpyxl.load_workbook(kb_path)

        # Construct header dictionary
        self.header={}
        for ws in self.kb.worksheets:
            self.header[ws.title.lower()]={}
            self.header[ws.title]={}
            for columnIdx in range(1,50):
                if ws.cell(row=1, column=columnIdx).value is not None:
                    self.header[ws.title.lower()][ws.cell(row=1, column=columnIdx).value.lower()] = columnIdx
                    self.header[ws.title][ws.cell(row=1, column=columnIdx).value] = columnIdx

    def translateRxns(self):
        """ Parses information from the core's Reaction sheet and inserts them to the appropiate field(s) in the KB structure. """

        coreRxns = self.core['Reactions']
        kbRxns   = self.kb['Reactions']

        for rowIdx in range(3, coreRxns.max_row+1):
            print(rowIdx)
            reactionId = coreRxns.cell(column=1, row=rowIdx-1).value

            kbRxns.cell(column=1, row=rowIdx-1).value = reactionId # ID
            kbRxns.cell(column=2, row=rowIdx-1).value = coreRxns.cell(column=2, row=rowIdx).value # Name
            #kbRxns.cell(column=3, row=rowIdx-1).value = coreRxns.cell(column=3, row=rowIdx).value # Synonyms

            # Type
            rxnType = coreRxns.cell(column=5, row=rowIdx).value # TYPE
            if rxnType is None:
                kbRxns.cell(column=4, row=rowIdx-1).value = 'Uncategorized'
            elif rxnType in wc_kb.core.ReactionType:
                kbRxns.cell(column=4, row=rowIdx-1).value = rxnType
            else:
                kbRxns.cell(column=4, row=rowIdx-1).value = 'misc'

            #kbRxns.cell(column=self.header['reactions']['submodel'], row=rowIdx-1).value = coreRxns.cell(column=6, row=rowIdx).value # Participants
            kbRxns.cell(column=self.header['reactions']['participants'], row=rowIdx-1).value = coreRxns.cell(column=6, row=rowIdx).value # Participants

            # Reversible
            reversible = coreRxns.cell(column=7, row=rowIdx).value #Reversibility
            if reversible == 'f':
                kbRxns.cell(column=self.header['reactions']['reversible'], row=rowIdx-1).value = False
            elif reversible == 'r':
                kbRxns.cell(column=self.header['reactions']['reversible'], row=rowIdx-1).value = True

            # Parameters
            parameterEntry = self.createParameters(
                name  = 'V_max of {}'.format(reactionId),
                value = coreRxns.cell(column=29, row=rowIdx).value,
                units = coreRxns.cell(column=30, row=rowIdx).value)

            evidenceEntry = self.createEvidence(
                objectId = reactionId,
                property = 'V_max',
                values = coreRxns.cell(column=15, row=rowIdx).value,
                units = coreRxns.cell(column=30, row=rowIdx).value,
                comments = coreRxns.cell(column=31, row=rowIdx).value)

            if parameterEntry is not None:
                kbRxns.cell(column=self.header['reactions']['parameters'], row=rowIdx-1).value = parameterEntry['id']
                if evidenceEntry is not None:
                    self.insertCell(insertTo=parameterEntry, header='Evidence', value=evidenceEntry['id'])

            kbRxns.cell(column=self.header['reactions']['spontenaeous'], row=rowIdx-1).value = coreRxns.cell(column=22, row=rowIdx).value
            kbRxns.cell(column=self.header['reactions']['enzyme'], row=rowIdx-1).value = coreRxns.cell(column=12, row=rowIdx).value
            kbRxns.cell(column=self.header['reactions']['coenzymes'], row=rowIdx-1).value = coreRxns.cell(column=15, row=rowIdx).value
            kbRxns.cell(column=self.header['reactions']['database references'], row=rowIdx-1).value = self.parseDatabaseReferences(coreRxns.cell(column=4, row=rowIdx))
            kbRxns.cell(column=self.header['reactions']['comments'], row=rowIdx-1).value = coreRxns.cell(column=41, row=rowIdx).value
            kbRxns.cell(column=self.header['reactions']['references'], row=rowIdx-1).value = coreRxns.cell(column=42, row=rowIdx).value

    def translateRnas(self):
        """ Parses information from the core's Rna sheet and inserts them to the appropiate field(s) in the KB structure. """

        coreRnas = self.core['RNA']
        kbRnas   = self.kb['RNAs']

        for rowIdx in range(3, coreRnas.max_row+1):
            print(rowIdx)
            speciesTypeId = kbRnas.cell(column=1, row=rowIdx-1).value
            assert speciesTypeId == coreRnas.cell(column=1, row=rowIdx).value

            #MANUALLY COPIED: kbRnas.cell(column=1, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # ID
            #MANUALLY COPIED: kbRnas.cell(column=2, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # NAME
            #MANUALLY COPIED: kbRnas.cell(column=3, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # Synonyms
            #MANUALLY COPIED: kbRnas.cell(column=4, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # Type
            #MANUALLY COPIED: kbRnas.cell(column=5, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # TU
            #MANUALLY COPIED: kbRnas.cell(column=6, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # Genes
            #MANUALLY COPIED: kbRnas.cell(column=7, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # Coordinate
            #MANUALLY COPIED: kbRnas.cell(column=8, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # Length
            #MANUALLY COPIED: kbRnas.cell(column=9, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # Direction
            #MANUALLY COPIED: kbRnas.cell(column=10, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # DB refs
            #MANUALLY COPIED: kbRnas.cell(column=11, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # references
            #MANUALLY COPIED: kbRnas.cell(column=12, row=rowIdx-1).value = coreRnas.cell(column=, row=rowIdx).value # comments

            # Species properties
            speciesPropsEntry = self.createSpeciesTypeProperties(
                speciesTypeId = speciesTypeId,
                halfLife = coreRnas.cell(column=15, row=rowIdx).value,
                halfLifeUnits = 'min')

            if coreRnas.cell(column=17, row=rowIdx).value is not None:
                speciesPropsEvidenceList = self.delinateJSONs(coreRnas.cell(column=17, row=rowIdx).value)
                for evidence in speciesPropsEvidenceList:
                    if evidence['value'] == coreRnas.cell(column=15, row=rowIdx).value:
                        refs  = self.cleanJsonDump(''.join(evidence['references']))
                        expId = self.findExperiment(specie=evidence['species'], refs=refs)
                        if expId is None:
                            expEntry = self.createExperiment(species=evidence['species'], refs=refs)
                            expId = expEntry['id']

                evidenceEntry = self.createEvidence(
                    objectId = speciesTypeId,
                    property = 'half_life',
                    values = coreRnas.cell(column=15, row=rowIdx).value,
                    units = 'min',
                    experiment = expId)

            if speciesPropsEntry is not None:
                kbRnas.cell(column=self.header['RNAs']['Species properties'], row=rowIdx-1).value = speciesPropsEntry['id']
                if evidenceEntry is not None:
                    self.insertCell(insertTo=speciesPropsEntry, header='Evidence', value=evidenceEntry['id'])

            # Concentration
            concentrationEntry = self.createConcentration(
                speciesTypeId = speciesTypeId ,
                values = coreRnas.cell(column=12, row=rowIdx).value,
                units = 'molecules')

            concEvidenceList = self.delinateJSONs(coreRnas.cell(column=14, row=rowIdx).value)
            for evidence in concEvidenceList:
                if evidence['value'] == coreRnas.cell(column=12, row=rowIdx).value:

                    refs  = self.cleanJsonDump(''.join(evidence['references']))
                    expId = self.findExperiment(specie=evidence['species'], refs=refs)
                    if expId is None:
                        expEntry = self.createExperiment(species=evidence['species'], refs=refs)
                        expId = expEntry['id']

            evidenceEntry = self.createEvidence(
                objectId = speciesTypeId,
                property = 'concentration',
                values = coreRnas.cell(column=12, row=rowIdx).value,
                units = 'molecules',
                experiment = expId)

            if concentrationEntry is not None:
                kbRnas.cell(column=self.header['RNAs']['Concentration'], row=rowIdx-1).value = concentrationEntry['id']
                if evidenceEntry is not None:
                    self.insertCell(insertTo=concentrationEntry, header='Evidence', value=evidenceEntry['id'])

    def translateMetabolites(self):
        """ Parses information from the core's Metabolite sheet and inserts them to the appropiate field(s) in the KB structure. """

        coreMetas = self.core['Metabolites']
        kbMetas   = self.kb['Metabolites']

        for rowIdx in range(3, coreMetas.max_row+1):
            print(rowIdx)
            speciesTypeId = kbMetas.cell(column=1, row=rowIdx-1).value
            assert speciesTypeId == coreMetas.cell(column=1, row=rowIdx).value

            #MANUALLY COPIED: kbMetas.cell(column=1, row=rowIdx-1).value = coreMetas.cell(column=1, row=rowIdx).value # ID
            #MANUALLY COPIED: kbMetas.cell(column=2, row=rowIdx-1).value = coreMetas.cell(column=2, row=rowIdx).value # NAME
            #MANUALLY COPIED: kbMetas.cell(column=3, row=rowIdx-1).value = coreMetas.cell(column=5, row=rowIdx).value # Synonyms

            # Type
            metaType = coreMetas.cell(column=7, row=rowIdx).value # TYPE
            if metaType is None:
                kbMetas.cell(column=4, row=rowIdx-1).value = 'unknown'
            elif metaType not in wc_kb.core.MetaboliteSpeciesTypeType:
                kbMetas.cell(column=4, row=rowIdx-1).value = 'misc'
            elif metaType not in wc_kb.core.MetaboliteSpeciesTypeType:
                kbMetas.cell(column=4, row=rowIdx-1).value = metaType

            # Species properties
            speciesPropsEntry = self.createSpeciesTypeProperties(
                speciesTypeId = speciesTypeId,
                structure = coreMetas.cell(column=9, row=rowIdx).value)
            if speciesPropsEntry is not None:
                kbMetas.cell(column=self.header['Metabolites']['Species properties'], row=rowIdx-1).value = speciesPropsEntry['id']

            #MANUALLY COPIED: kbMetas.cell(column=8, row=rowIdx-1).value = coreMetas.cell(column=refsColumn, row=rowIdx).value # REFERENCES
            #MANUALLY COPIED: kbMetas.cell(column=9, row=rowIdx-1).value = coreMetas.cell(column=commentsColumn, row=rowIdx).value #comments
            kbMetas.cell(column=self.header['Metabolites']['Database references'], row=rowIdx-1).value = \
                self.parseDatabaseReferences(cell = coreMetas.cell(column=6, row=rowIdx)) # Add DB refs

            # Concentration
            if coreMetas.cell(column=18, row=rowIdx).value is not None:
                concDict = json.loads(coreMetas.cell(column=18, row=rowIdx).value)

                concentrationEntry = self.createConcentration(
                    speciesTypeId = speciesTypeId ,
                    values = concDict['concentration'],
                    units = 'mM')

                if 'evidence' not in concDict.keys():
                    continue

                for evidence in concDict['evidence']:
                    if evidence['value'] == concDict['concentration']:

                        refs  = self.cleanJsonDump(''.join(evidence['references']))
                        expId = self.findExperiment(specie=evidence['species'], refs=refs)
                        if expId is None:
                            expEntry = self.createExperiment(species=evidence['species'], refs=refs)
                            expId = expEntry['id']

                        try:
                            comments = evidence['comments']
                        except:
                            comments = None

                        evidenceEntry = self.createEvidence(
                            objectId = speciesTypeId,
                            property = 'concentration',
                            values = evidence['value'],
                            units = evidence['units'],
                            experiment = expId,
                            comments = comments)

                if concentrationEntry is not None:
                    kbMetas.cell(column=self.header['Metabolites']['Concentration'], row=rowIdx-1).value = concentrationEntry['id']
                    if evidenceEntry is not None:
                        self.insertCell(insertTo=concentrationEntry, header='Evidence', value=evidenceEntry['id'])

    def translateProteins(self):
        """ Parses information from the core's 'protein monomers' sheet and inserts them to the appropiate field(s) in the KB structure. """

        header=self.header
        coreProteins = self.core['Protein monomers']
        kbProteins = self.kb['Proteins']
        kbSpeciesProperties = self.kb['Species properties']
        kbConcentrations = self.kb['Concentrations']

        # Create experiment entries
        concExpEntry = self.createExperiment(
                              species='Mycoplasma Pneumoniae',
                              geneticVariant='M129',
                              externalMedia='Hayflick',
                              temperature=37,
                              refs='PUB_0932',
                              comments='Concentrations available at multiple time points, see the Comments in the evidence field')
        hlExpEntry = self.createExperiment(
                              species='Mycoplasma Pneumoniae',
                              geneticVariant='M129',
                              externalMedia='Hayflick',
                              temperature=37,
                              refs='PUB_0956')
        methionineExpEntry = self.createExperiment(
                              species='Mycoplasma Pneumoniae',
                              geneticVariant='M129',
                              refs='PUB_0937')

        for rowIdx in range(3, coreProteins.max_row+1):
            print(rowIdx)

            # Check if row is correct
            speciesTypeId = kbProteins.cell(column=1, row=rowIdx-1).value
            assert speciesTypeId == coreProteins.cell(column=1, row=rowIdx).value

            #MANUALLY COPIED: kbProteins.cell(column=1, row=rowIdx-1).value = coreProteins.cell(column=1, row=rowIdx).value # ID
            #MANUALLY COPIED: kbProteins.cell(column=2, row=rowIdx-1).value = coreProteins.cell(column=2, row=rowIdx).value # Name
            #MANUALLY COPIED: kbProteins.cell(column=3, row=rowIdx-1).value = coreProteins.cell(column=3, row=rowIdx).value # Synonyms

            # Types
            proteinType = coreProteins.cell(column=5, row=rowIdx).value
            if proteinType is None:
                kbProteins.cell(column=4, row=rowIdx-1).value = 'uncategorized'
            else:
                kbProteins.cell(column=4, row=rowIdx-1).value = coreProteins.cell(column=5, row=rowIdx).value

            # Species properties
            speciesPropsEntry = self.createSpeciesTypeProperties(
                speciesTypeId = speciesTypeId,
                halfLife = coreProteins.cell(column=27, row=rowIdx).value,
                domains= coreProteins.cell(column=14, row=rowIdx).value,
                prostheticGroups = coreProteins.cell(column=15, row=rowIdx).value)
            evidenceEntry = self.createEvidence(
                objectId = speciesTypeId,
                property = 'half life',
                values = coreProteins.cell(column=27, row=rowIdx).value,
                units = 'min',
                experiment = hlExpEntry['id'])
            if speciesPropsEntry is not None:
                kbProteins.cell(column=self.header['Proteins']['Species properties'], row=rowIdx-1).value = speciesPropsEntry['id']
                if evidenceEntry is not None:
                    self.insertCell(insertTo=speciesPropsEntry, header='Evidence', value=evidenceEntry['id'])

            # Concentration
            concentrationEntry = self.createConcentration(
                speciesTypeId = speciesTypeId ,
                values = coreProteins.cell(column=22, row=rowIdx).value,
                units = 'molecules')
            evidenceEntry = self.createEvidence(
                objectId = speciesTypeId,
                property = 'concentration',
                values = coreProteins.cell(column=22, row=rowIdx).value,
                units = 'molecules',
                experiment = concExpEntry['id'],
                comments=coreProteins.cell(column=23, row=rowIdx).value)
            if concentrationEntry is not None:
                kbProteins.cell(column=self.header['Proteins']['Concentration'], row=rowIdx-1).value = concentrationEntry['id']
                if evidenceEntry is not None:
                    self.insertCell(insertTo=concentrationEntry, header='Evidence', value=evidenceEntry['id'])

            #MANUALLY COPIED: kbProteins.cell(column=5, row=rowIdx-1).value = coreProteins.cell(column=6, row=rowIdx).value # Gene
            kbProteins.cell(column=self.header['Proteins']['Localization'], row=rowIdx-1).value = coreProteins.cell(column=9, row=rowIdx).value.replace('t', '') # localization: treat trans- membrane
            #MANUALLY COPIED: kbProteins.cell(column=9, row=rowIdx-1).value = coreProteins.cell(column=6, row=rowIdx).value # Translation rates
            #MANUALLY COPIED: kbProteins.cell(column=10, row=rowIdx-1).value = coreProteins.cell(column=6, row=rowIdx).value # Translation rate units

            evidenceEntry = self.createEvidence(
                objectId = speciesTypeId,
                property = 'translationRate',
                values = coreProteins.cell(column=24, row=rowIdx).value,
                units = '1/s/mRNA',
                experiment = hlExpEntry['id']) #measured in the same experiment as half lifes
            if evidenceEntry is not None:
                proteinEntry ={'ws':'Proteins', 'id':speciesTypeId, 'row':rowIdx-1}
                self.insertCell(insertTo=proteinEntry, header='Evidence', value=evidenceEntry['id'])

            #MANUALLY COPIED: kbProteins.cell(column=11, row=rowIdx-1).value = coreProteins.cell(column=6, row=rowIdx).value # Is methionine cleaved?

            evidenceEntry = self.createEvidence(
                objectId = speciesTypeId,
                property = 'Is methionine cleaved',
                values = coreProteins.cell(column=7, row=rowIdx).value,
                units = 'dimensionless',
                experiment = methionineExpEntry['id'])
            if evidenceEntry is not None:
                proteinEntry ={'ws':'Proteins', 'id':speciesTypeId, 'row':rowIdx-1}
                self.insertCell(insertTo=proteinEntry, header='Evidence', value=evidenceEntry['id'])

            #MANUALLY COPIED: kbProteins.cell(column=7, row=rowIdx-1).value = coreProteins.cell(column=10, row=rowIdx).value # signal sequence type
            #MANUALLY COPIED: kbProteins.cell(column=8, row=rowIdx-1).value = coreProteins.cell(column=11, row=rowIdx).value # signal sequence localization
            #MANUALLY COPIED: kbProteins.cell(column=9, row=rowIdx-1).value = coreProteins.cell(column=12, row=rowIdx).value # signal sequence length
            #MANUALLY COPIED: kbProteins.cell(column=10, row=rowIdx-1).value = coreProteins.cell(column=16, row=rowIdx).value # DNA footprint length
            #MANUALLY COPIED: kbProteins.cell(column=11, row=rowIdx-1).value = coreProteins.cell(column=17, row=rowIdx).value # DNA footprint biding
            #MANUALLY COPIED: kbProteins.cell(column=16, row=rowIdx-1).value = coreProteins.cell(column=33, row=rowIdx).value # References
            #MANUALLY COPIED: kbProteins.cell(column=17, row=rowIdx-1).value = coreProteins.cell(column=34, row=rowIdx).value # Comments

            kbProteins.cell(column=self.header['Proteins']['Database references'], row=rowIdx-1).value = \
                self.parseDatabaseReferences(cell = coreProteins.cell(column=4, row=rowIdx))

    def translateChromosomeFeatures(self):
        """ Parses information from the core's 'Chromosome feature' sheet and inserts them to the appropiate field(s) in the KB structure. """

        coreChromFeats = self.core['Chromosome features']
        kbChromFeats = self.kb['Chromosome features']

        for rowIdx in range(3, coreChromFeats.max_row+1):
            print(rowIdx)

            # Check if row is correct
            chromeFeatId = kbChromFeats.cell(column=1, row=rowIdx-1).value
            assert chromeFeatId == coreChromFeats.cell(column=1, row=rowIdx).value

            chromFeatDict = coreChromFeats.cell(column=27, row=rowIdx).value
            if chromFeatDict is None:
                continue

            evidenceEntry = self.createEvidence(
                objectId = chromeFeatId,
                property = 'intensity',
                values = chromFeatDict['value'],
                units = chromFeatDict['units'],
                experiment = 'EXP0001',
                comments = chromFeatDict['comments'],
                refs = 'PUB_0959')

            kbChromFeats.cell(column=self.header['Chromosome Features']['Evidence'], row=rowIdx-1).value = evidenceEntry['id']

    """ Create nested entries """
    def createParameters(self, name=None, value=None, units=None, evidence=None):
        if value is None:
            return None

        wsName = 'Parameters'
        nextRow = self.kb[wsName].max_row+1
        parameterId = 'parameter_{}'.format(nextRow-1)

        self.kb[wsName].cell(column=self.header['parameters']['id'], row=nextRow).value = parameterId
        self.kb[wsName].cell(column=self.header['parameters']['name'], row=nextRow).value = name
        self.kb[wsName].cell(column=self.header['parameters']['value'], row=nextRow).value = value
        self.kb[wsName].cell(column=self.header['parameters']['units'], row=nextRow).value = units
        self.kb[wsName].cell(column=self.header['parameters']['evidence'], row=nextRow).value = evidence

        entry ={'ws':'Parameters', 'id':parameterId, 'row':nextRow}
        return entry

    def createSpeciesTypeProperties(self, speciesTypeId, structure=None, halfLife=None, halfLifeUnits='min', domains=None, prostheticGroups=None, evidence=None, dbRefs=None, refs=None, comments=None):
        inputs = [structure, halfLife, domains, prostheticGroups]
        if all(input is None for input in inputs):
            return None

        wsName = 'Species properties'
        nextRow = self.kb[wsName].max_row+1
        speciesPropId = 'props_{}'.format(speciesTypeId)

        self.kb[wsName].cell(column=self.header['species properties']['id'], row=nextRow).value = speciesPropId
        self.kb[wsName].cell(column=self.header['species properties']['structure'], row=nextRow).value = structure
        self.kb[wsName].cell(column=self.header['species properties']['half life'], row=nextRow).value = halfLife
        self.kb[wsName].cell(column=self.header['species properties']['half life units'], row=nextRow).value = halfLifeUnits
        self.kb[wsName].cell(column=self.header['species properties']['domains'], row=nextRow).value = domains
        self.kb[wsName].cell(column=self.header['species properties']['prosthetic groups'], row=nextRow).value = prostheticGroups
        self.kb[wsName].cell(column=self.header['species properties']['evidence'], row=nextRow).value = evidence
        self.kb[wsName].cell(column=self.header['species properties']['database references'], row=nextRow).value = dbRefs
        self.kb[wsName].cell(column=self.header['species properties']['references'], row=nextRow).value = refs
        self.kb[wsName].cell(column=self.header['species properties']['comments'], row=nextRow).value = comments

        entry ={'ws':wsName, 'id':speciesPropId, 'row':nextRow}
        return entry

    def createConcentration(self, speciesTypeId, compartment='c', values=None, units=None, evidence=None, dbRefs=None, refs=None, comments=None):
        assert compartment in ['c', 'e', 'm']
        if values is None:
            return None

        wsName = 'Concentrations'
        nextRow = self.kb[wsName].max_row+1
        concentrationId = 'conc_{}_{}'.format(speciesTypeId, compartment)

        self.kb[wsName].cell(column=self.header['concentrations']['id'], row=nextRow).value = concentrationId
        self.kb[wsName].cell(column=self.header['concentrations']['species'], row=nextRow).value = '{}[{}]'.format(speciesTypeId, compartment)
        self.kb[wsName].cell(column=self.header['concentrations']['values'], row=nextRow).value = values
        self.kb[wsName].cell(column=self.header['concentrations']['units'], row=nextRow).value = units
        self.kb[wsName].cell(column=self.header['concentrations']['evidence'], row=nextRow).value = evidence
        self.kb[wsName].cell(column=self.header['concentrations']['database references'], row=nextRow).value = dbRefs
        self.kb[wsName].cell(column=self.header['concentrations']['references'], row=nextRow).value = refs
        self.kb[wsName].cell(column=self.header['concentrations']['comments'], row=nextRow).value = comments

        entry ={'ws':'Concentrations', 'id':concentrationId, 'row':nextRow}
        return entry

    def createEvidence(self, objectId, property=None, values=None, units=None, experiment=None, dbRefs=None, refs=None, comments=None):
        inputs = [property, values, units, experiment, dbRefs, refs, comments]
        #if all(input is None for input in inputs):
        if values is None:
            return None

        wsName = 'Evidence'
        nextRow = self.kb[wsName].max_row+1
        eviId   = 'EVI{}'.format(str(nextRow-1).zfill(4))

        self.kb[wsName].cell(column=self.header['evidence']['id'], row=nextRow).value = eviId
        self.kb[wsName].cell(column=self.header['evidence']['cell'], row=nextRow).value = 'cell'
        self.kb[wsName].cell(column=self.header['evidence']['object'], row=nextRow).value = objectId
        self.kb[wsName].cell(column=self.header['evidence']['property'], row=nextRow).value = property
        self.kb[wsName].cell(column=self.header['evidence']['values'], row=nextRow).value = values
        self.kb[wsName].cell(column=self.header['evidence']['units'], row=nextRow).value = units
        self.kb[wsName].cell(column=self.header['evidence']['experiment'], row=nextRow).value = experiment
        self.kb[wsName].cell(column=self.header['evidence']['database references'], row=nextRow).value = dbRefs
        self.kb[wsName].cell(column=self.header['evidence']['comments'], row=nextRow).value = comments

        entry ={'ws':'Evidence', 'id':eviId, 'row':nextRow}
        return entry

    def createExperiment(self, experimentDesign=None, measurmentTechnology=None, analysisType=None, species=None, geneticVariant=None, externalMedia=None, temperature=None, temperatureUnits='C', ph=None, dbRefs=None, refs=None, comments=None):
        wsName = 'Experiment'
        nextRow = self.kb[wsName].max_row+1
        expId   = 'EXP{}'.format(str(nextRow-1).zfill(4))

        self.kb[wsName].cell(column=self.header['experiment']['id'], row=nextRow).value = expId
        self.kb[wsName].cell(column=self.header['experiment']['experiment design'], row=nextRow).value = experimentDesign
        self.kb[wsName].cell(column=self.header['experiment']['measurment technology'], row=nextRow).value = measurmentTechnology
        self.kb[wsName].cell(column=self.header['experiment']['analysis type'], row=nextRow).value = analysisType
        self.kb[wsName].cell(column=self.header['experiment']['species'], row=nextRow).value = species
        self.kb[wsName].cell(column=self.header['experiment']['genetic variant'], row=nextRow).value = geneticVariant
        self.kb[wsName].cell(column=self.header['experiment']['external media'], row=nextRow).value = externalMedia
        self.kb[wsName].cell(column=self.header['experiment']['temperature'], row=nextRow).value = temperature
        self.kb[wsName].cell(column=self.header['experiment']['temperature units'], row=nextRow).value = temperatureUnits
        self.kb[wsName].cell(column=self.header['experiment']['ph'], row=nextRow).value = ph
        self.kb[wsName].cell(column=self.header['experiment']['database references'], row=nextRow).value = dbRefs
        self.kb[wsName].cell(column=self.header['experiment']['references'], row=nextRow).value = refs
        self.kb[wsName].cell(column=self.header['experiment']['comments'], row=nextRow).value = comments

        entry ={'ws':'Experiment', 'id':expId, 'row':nextRow}
        return entry

    """ Auxiliary functions """
    def insertCell(self, insertTo, header, value, concatenate=True):
        wsName = insertTo['ws']
        rowIdx = insertTo['row']

        if  concatenate is True:
            self.concatenateId(self.kb[wsName].cell(column=self.header[wsName][header], row=rowIdx), value)
        else:
            self.kb[wsName].cell(column=self.header[wsName][header], row=rowIdx).value =value

    @staticmethod
    def concatenateId(cell, entry):
        if entry is None:
            return

        if cell.value is None:
            cell.value = '' +  entry
        else:
            cell.value = cell.value + ', ' +  entry

    def parseDatabaseReferences(self, cell):
        if cell.value is None:
            return None

        jsons = self.delinateJSONs(cell.value)

        if jsons is None:
            return None
        assert isinstance(jsons, list)
        assert len(jsons) > 0

        dbRefIds = ''
        for json in jsons:

            if json['source']=='URL':
                pass #dbRefId = "{}:{}".format(json['source'].lower(), str(nextRow).zfill(4))
            else:
                dbRefId = "{}:{}".format(str(json['source']).lower(), str(json['xid']))

            dbRefId = dbRefId.replace('-','_')
            dbRefIds += dbRefId + ', '

        return dbRefIds[:-2]

    def findExperiment(self, specie=None, refs=None):

        wsName = 'Experiment'
        nextRow = self.kb[wsName].max_row+1

        for rowIdx in range(2,nextRow):
             if self.kb[wsName].cell(row=rowIdx, column=self.header['experiment']['species']).value == specie and \
                self.kb[wsName].cell(row=rowIdx, column=self.header['experiment']['references']).value == refs:
                experimentId = self.kb[wsName].cell(row=rowIdx, column=1).value
                return experimentId

        return None

    def idExists(self, wsName, id):
        """ Check if ID exists within worksheet """

        for row in range(1, self.kb[wsName].max_row+1):
            if self.kb[wsName].cell(column=1, row=row).value == id:
                return True

        return False

    @staticmethod
    def delinateJSONs(fieldStr):
        """ Detects and return individual JSON strings within fieldStr """

        if fieldStr is None or fieldStr=='':
            return None

        jsons = []
        jsonStrLen = 0
        matches = re.findall(r'{.*?}', fieldStr)
        for match in matches:
            jsonStrLen += len(match)
            try:
                jsons.append(json.loads(match))
            except:
                raise Exception('\nCan not parse JSONs from line: \n\t {}'.format(fieldStr))

        # Check if whole fieldStr has been parsed to JSONs; assumes JSONS are separated as "}, {"
        if jsonStrLen+(len(matches)-1)*2 != len(fieldStr):
            warnings.warn('\nString contains non-JSON substrings: \n\t {}'.format(fieldStr))

        # Return None if only non-JSON substrings are found
        if jsons ==[]:
            return None

        return jsons

    @staticmethod
    def cleanJsonDump(fieldStr):
        fieldStr = fieldStr.replace('[', '')
        fieldStr = fieldStr.replace(']', '')
        fieldStr = fieldStr.replace('"', '')
        return fieldStr
